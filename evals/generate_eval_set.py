from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook


DEFAULT_DATASET_PATH = "evals/评测集.xlsx"
EXCEL_HEADERS = [
    "编号",
    "original_question",
    "rewritten_question",
    "intent",
    "difficulty",
    "user_profile",
    "current_plan",
    "execution_result",
    "history_high_freq_questions",
    "reference_top3",
    "evaluation_focus",
    "scenario_family",
    "region",
    "intent_mode",
]


TOTAL_CASES = 120
REGIONS = ["泰国北区", "泰国南区", "泰国东区", "泰国西区", "曼谷核心区", "东北片区"]
TIME_WINDOWS = ["昨晚 20:00-22:00", "今天 08:00-10:00", "近 24 小时", "近 3 天", "本周高峰时段", "凌晨 01:00-03:00"]
THRESHOLDS = ["10%", "12%", "15%"]
DIFFICULTIES = (["easy"] * 30) + (["medium"] * 66) + (["hard"] * 24)

ROLES = [
    "网络监控工程师",
    "网络规划工程师",
    "网络维护工程师",
]
GEO_TERMS = [
    "地理术语习惯：北区 = 泰国清迈府及周边标准区域",
    "地理术语习惯：南区 = 泰国宋卡府及周边标准区域",
    "地理术语习惯：东区 = 泰国春武里府及周边标准区域",
    "地理术语习惯：西区 = 泰国北碧府及周边标准区域",
]
NETWORK_TERMS = [
    "网络术语习惯：3号设备 = CSW-03 核心交换机",
    "网络术语习惯：4号设备 = UPF-04 用户面网元",
    "网络术语习惯：A池 = 曼谷核心资源池 A",
    "网络术语习惯：B链 = 北区回传主承载链路 B",
]
BUSINESS_TERMS = [
    "业务术语习惯：趴网状态 = 黄金指标监控任务",
    "业务术语习惯：保活任务 = 网元连通性巡检任务",
    "业务术语习惯：抢通动作 = 主备倒换与容灾恢复流程",
    "业务术语习惯：健康度日报 = 关键指标天报任务",
]

POOLS = {
    "知识问答": {
        "sites": ["BKK-AMF-011", "CNX-SMF-008", "HKT-UPF-022", "NRT-IMS-105", "HKT-OLT-023", "CNX-OLT-014"],
        "nodes": ["AMF", "SMF", "UPF", "IMS", "OLT", "ONU"],
        "metrics": ["告警级别", "恢复时长", "关联 KPI 波动", "注册成功率", "健康度报表", "巡检通过率"],
        "alarms": ["SCTP 重传告警", "接口超时告警", "LOS 告警", "资源池切换告警", "CPU 高负载告警", "PON 口拥塞告警"],
        "codes": ["15", "11", "13", "98", "111", "301"],
        "services": ["5G 注册", "会话建立", "黄金指标监控任务", "保活任务", "健康度日报", "家宽接入"],
        "objects": ["网元", "资源池", "任务流程", "告警对象", "巡检任务", "报表对象"],
    },
    "网络监控": {
        "sites": ["CNX-5G-001", "BKK-LTE-118", "AYT-NR-021", "BKK-AMF-011", "KBI-MW-009", "HKT-PTN-004"],
        "nodes": ["gNodeB", "AMF", "UPF", "IPRAN", "PTN", "监控任务"],
        "metrics": ["注册成功率", "掉话率", "时延", "抖动", "历史比对结果", "扩容阈值逼近度"],
        "alarms": ["链路抖动告警", "高干扰告警", "SCTP 重传告警", "端口拥塞告警", "CPU 高负载告警", "小区退服告警"],
        "codes": ["15", "11", "13", "98", "111", "384"],
        "services": ["历史查询", "历史比对", "横向比对", "阈值比对", "资源趋势预测", "关键指标天报"],
        "objects": ["区域", "站点", "小区", "链路", "资源池", "扩容对象"],
    },
    "故障处理": {
        "sites": ["CNX-5G-001", "BKK-LTE-118", "AYT-NR-021", "BKK-AMF-011", "NRT-IMS-105", "HKT-OLT-023"],
        "nodes": ["gNodeB", "AMF", "SMF", "IMS", "OLT", "IPRAN"],
        "metrics": ["注册失败率", "掉话率", "时延", "抖动", "恢复时长", "残留告警数"],
        "alarms": ["SCTP 重传告警", "链路抖动告警", "高干扰告警", "LOS 告警", "接口超时告警", "资源池切换告警"],
        "codes": ["15", "11", "13", "98", "111", "301"],
        "services": ["故障诊断", "抢通恢复", "值守监控", "告警工单", "KPI 工单", "容灾切换"],
        "objects": ["异常区域", "高影响对象", "工单对象", "恢复对象", "值守对象", "业务对象"],
    },
}


@dataclass(frozen=True)
class Blueprint:
    key: str
    intent: str
    repeats: int
    question_template: str
    rewritten_template: str
    plan_steps: list[str]
    findings: list[str]
    uncertainties: list[str]
    next_questions: list[str]
    history_templates: list[str]
    evaluation_focus: list[str]


BLUEPRINTS = [
    Blueprint("alarm_knowledge", "知识问答", 10, "{region} 的 {alarm} 想讲清楚原理、影响和处理思路，下一步应该怎么问？", "回答 {alarm} 的知识问题，并补充解释、定位和恢复思路", ["澄清定义与级别", "补充影响对象与原因", "说明定位步骤", "补充恢复与清除机制", "给出使用指导"], ["已整理 {alarm} 的基础描述、影响对象和常见触发场景，并确认其主要关联 {node} 与 {service}。", "当前执行结果已经覆盖告警级别、上报主体和部分定位线索，并识别出该告警常与 {metrics} 变化一起出现。"], ["但还没有把内容落到“现场先看什么、后看什么”的使用顺序上。", "不过目前仍缺少与关键异常信号、清除机制或根因告警之间的明确对应关系。"], ["是否继续补充 {alarm} 在现场排查时应该优先查看哪些对象和信号？", "要不要把 {alarm} 的可能原因按触发阶段和影响范围继续拆开说明？", "是否继续补充 {alarm} 与关键异常信号或恢复动作之间的对应关系？", "要不要把当前知识说明进一步落到可执行的处理步骤上？"], ["{alarm} 是什么？", "{alarm} 为什么会出现？", "{alarm} 怎么处理？", "{alarm} 清除机制是什么？"], ["next_step", "knowledge_grounding"]),
    Blueprint("core_principle_knowledge", "知识问答", 10, "想从现网运维视角理解 {node} 在 {service} 里的作用，下一步应该怎么问？", "回答 {node} 的原理与功能知识问题，并补充 {service} 场景中的定位思路", ["澄清职责边界", "说明业务流程位置", "补充与其他网元关系", "说明定位思路", "补充现网使用指导"], ["已说明 {node} 在 {service} 里的基本职责，并列出了与相关对象的交互位置。", "系统已把 {node} 的概念、功能边界和关键观测点做了基础整合，并映射到典型流程。"], ["但还没有把这些知识转成“遇到问题时先看哪里”的使用顺序。", "不过现有结果仍缺少与关键指标、告警或流程阶段之间的联系说明。"], ["是否继续说明 {node} 在实际任务中应先看哪些信号或结果来判断是否异常？", "要不要把 {node} 的职责差异进一步落到 {service} 的关键流程步骤上？", "是否继续补充 {node} 与相关对象在定位问题时的观察顺序？", "要不要把当前知识说明进一步转成一套可执行的使用指导？"], ["{node} 是做什么的？", "{node} 在 {service} 中起什么作用？", "{node} 的原理是什么？", "{node} 故障时一般看什么？"], ["next_step", "knowledge_to_action"]),
    Blueprint("operation_guide_knowledge", "知识问答", 10, "想系统了解 {service} 怎么用，最好能讲清楚操作方法和流程步骤，下一步应该怎么问？", "回答 {service} 的操作方法与流程步骤知识问题，并补充使用指导", ["说明入口和适用前提", "梳理关键步骤", "补充结果查看方式", "说明常见误区", "给出使用建议"], ["已整理 {service} 的入口、主要步骤和基础结果查看方式，并对齐到 {objects} 的使用语境。", "当前执行结果已经覆盖标准操作步骤、部分注意事项和结果解释。"], ["但还没有把不同场景下“先做什么、后做什么”的顺序讲清楚。", "不过目前仍缺少对限制条件、常见误区和结果判定标准的补充。"], ["是否继续把 {service} 的操作方法补成一套明确的步骤顺序？", "要不要继续补充 {service} 在不同场景下的适用前提和限制条件？", "是否继续说明查看结果后该如何判断是否正常或需要继续处理？", "要不要把当前说明进一步补成适合一线使用的操作指导？"], ["{service} 怎么用？", "{service} 的操作步骤是什么？", "{service} 的结果怎么看？", "{service} 有什么限制？"], ["next_step", "workflow_guidance"]),
    Blueprint("private_knowledge", "知识问答", 10, "现在有一条客户私有知识相关的问题，围绕 {objects} 的使用习惯和口径还不够清楚，下一步应该怎么问？", "回答客户私有知识问题，并补充口径解释、使用边界和落地指导", ["澄清私有术语和口径", "确认适用对象和前提", "补充结果解释", "说明限制条件", "给出落地使用建议"], ["当前已整理出与 {objects} 相关的私有口径、常见表达和基础使用说明，并对齐到 {service} 场景。", "系统已汇总客户私有知识中的关键术语、适用对象和常见解释，但还停留在基础说明层面。"], ["但还没有把私有口径进一步落到“实际任务中怎么判断、怎么使用”的层面。", "不过目前仍缺少对边界条件、例外场景和结果判定方式的补充。"], ["是否继续补充客户私有口径在实际任务中的判断标准和使用边界？", "要不要把私有术语和标准对象、标准任务的对应关系继续讲清楚？", "是否继续说明在不同场景下应用这条私有知识时应先看什么、后看什么？", "要不要把当前说明进一步转成一线可直接使用的落地指导？"], ["客户私有知识怎么理解？", "这条私有口径适用于哪些对象？", "私有术语对应什么标准对象？", "这条知识下一步怎么用？"], ["next_step", "private_knowledge"]),
    Blueprint("monitor_kpi", "网络监控", 10, "想继续看 {region} 的 {metrics}，已经拿到初步结果，下一步问题怎么设计？", "围绕 {region} 的 {metrics} 查询结果，继续设计下一步监控分析问题", ["确认查询范围与口径", "查看当前值与历史值", "判断异常", "确定下钻对象", "补充比对策略"], ["当前已完成 {metrics} 的基础查询，并确认异常主要集中在 {region} 的部分 {objects} 上。", "系统已拉取 {service} 相关结果，发现局部对象波动明显高于基线。"], ["但还没有决定下一步应优先按时间、对象还是结果差异继续拆分。", "不过目前仍不清楚应先做历史比对、横向比对还是阈值过滤。"], ["是否继续按 {objects} 维度下钻，确认哪些对象贡献了主要异常？", "要不要把 {metrics} 按历史窗口和当前窗口继续对齐，确认异常从什么时候开始扩大？", "是否继续按阈值或异常幅度筛选对象，只保留最值得继续分析的部分？", "要不要先确定下一步是做历史比对、横向比对还是对象拆分？"], ["{metrics} 怎么查？", "{region} 的 {metrics} 当前值是多少？", "{metrics} 异常对象有哪些？", "{metrics} 下一步怎么分析？"], ["next_step", "monitoring_fit"]),
    Blueprint("monitor_status", "网络监控", 10, "查了 {objects} 的状态，只拿到初步结果，下一步还应该问什么？", "围绕 {objects} 状态查询结果继续设计下一步问题，以确认状态异常范围和判断标准", ["确认查询对象", "查看当前状态", "识别异常对象", "确定确认维度", "补充比对动作"], ["当前已完成 {objects} 的基础状态查询，发现部分对象存在异常或不稳定状态。", "系统已返回 {node} 的状态信息，并识别出少量异常对象或状态切换频繁的对象。"], ["但还没有继续确认异常是否集中在同一批对象或同一时间窗口。", "不过目前仍缺少用来判断“状态异常是否成立”的对比基线。"], ["是否继续按对象或时间窗口确认异常状态是否持续存在？", "要不要把当前状态结果与历史状态或基线继续做对比？", "是否继续筛选真正异常的对象，避免把短时抖动也当成问题？", "要不要进一步确认当前状态异常是否已经影响到相关任务或结果？"], ["{objects} 状态怎么查？", "{node} 当前状态是什么？", "哪些对象状态异常？", "状态异常后下一步怎么确认？"], ["next_step", "status_query_fit"]),
    Blueprint("monitor_report_compare", "网络监控", 10, "现在已经生成 {service} 的结果，但还不知道怎么继续定位异常重点，下一步应该问什么？", "围绕 {service} 的报表或对比结果继续设计下一步问题，定位异常重点和后续分析路径", ["确认统计周期和对比口径", "查看关键变化", "识别异常对象", "确定后续分析对象", "补充时间或对象维度分析"], ["当前已生成 {service} 的结果，并发现部分对象在最近周期内波动明显。", "系统已给出对比结果，并识别出少量对象在阈值或历史比对中偏离较大。"], ["但还没有继续按对象、时间或变化程度做更细的拆分。", "不过目前仍不清楚下一步是优先看趋势变化、异常排名还是阈值筛选。"], ["是否继续按异常对象做下一步下钻？", "要不要把结果按时间段继续拆开，确认异常是否集中在某些窗口？", "是否继续按变化幅度或阈值过滤对象，优先分析真正偏离基线的部分？", "要不要先确定下一步是看趋势变化、异常对象还是恢复情况？"], ["报表怎么看？", "哪些对象异常？", "差异来源怎么继续分析？", "下一步怎么下钻？"], ["next_step", "comparison_fit"]),
    Blueprint("resource_planning", "网络监控", 10, "当前在做 {service}，已经有一轮资源监控和统计分析结果了，下一步问题怎么设计更合理？", "围绕 {service} 的资源监控与趋势分析结果继续设计下一步问题，定位扩容对象与判断依据", ["确认资源对象和统计口径", "查看历史趋势与当前负荷", "识别高风险资源", "判断扩容优先级", "补充规划证据"], ["当前已完成 {objects} 的资源监控与趋势分析，确认部分对象在忙时段接近扩容阈值。", "系统已给出历史趋势、当前负荷和初步预测结果，并识别出少量高风险资源对象。"], ["但还没有继续按资源对象、时间段或增长速度进一步收敛范围。", "不过目前仍不清楚下一步应优先看历史趋势、当前负荷还是阈值逼近程度。"], ["是否继续按资源对象维度下钻，确认哪些对象最接近扩容阈值？", "要不要把历史趋势和当前忙时结果继续对齐，确认风险是持续存在还是短时波动？", "是否继续按阈值逼近程度筛选对象，优先保留最需要纳入 EXO 规划的部分？", "要不要先明确下一步是继续看增长趋势、当前负荷还是扩容优先级？"], ["资源趋势怎么看？", "哪些对象需要扩容？", "EXO 规划下一步怎么做？", "扩容优先级如何判断？"], ["next_step", "planning_fit"]),
    Blueprint("fault_diagnosis", "故障处理", 10, "{region} 的 {service} 出现异常，已经拿到初步执行结果，下一步应该问什么？", "围绕 {region} 的异常诊断结果继续设计下一步问题，补范围、补信号、补证据", ["确认影响范围", "核查关键告警和异常信号", "对齐时间窗口", "判断下钻对象", "补充证据链"], ["当前已完成影响范围确认和关键信号初查，确认异常主要落在 {region} 的部分 {objects} 上，同时伴随 {alarm}。", "系统已核查第一轮执行结果，发现 {metrics} 在故障时段明显偏离基线，并与 {code} 或关键异常信号相关。"], ["但还没有继续把关键信号和故障开始时刻完整对齐，也没有补齐交叉验证证据。", "不过目前仍不清楚异常是否集中在同一批对象，还是由多个方向共同触发。"], ["是否继续按 {objects} 维度确认异常是否集中在同一批对象？", "要不要把关键异常信号和故障开始时刻继续对齐，确认变化先后关系？", "是否继续补充另一侧结果或信号做交叉验证，帮助缩小可能原因？", "要不要先筛出最值得继续处理的高影响对象，再做下一步下钻？"], ["{service} 怎么排查？", "当前异常范围有多大？", "关键异常信号是什么？", "下一步该怎么继续诊断？"], ["next_step", "fault_diagnosis"]),
    Blueprint("fault_recovery", "故障处理", 10, "当前故障已经进入 {service} 阶段，但还不确定下一步抢通动作怎么问更合适。", "围绕 {service} 的抢通恢复结果继续设计下一步问题，确认恢复对象、动作顺序和恢复状态", ["确认需恢复的关键对象", "核查恢复动作结果", "判断是否需要继续倒换或迁移", "确认恢复后指标变化", "补充残留问题确认"], ["当前已完成第一步 {service} 动作，确认部分 {objects} 已恢复，但仍有残留影响对象存在。", "系统已执行初步恢复动作，并观察到 {metrics} 有所改善，但未完全回到基线水平。"], ["但还没有继续确认是否需要第二步恢复动作，或是否应先观察恢复后的稳定性。", "不过目前仍不清楚恢复是否覆盖了全部高影响对象，还是只恢复了局部。"], ["是否继续确认当前恢复动作是否已经覆盖全部高影响对象？", "要不要把恢复前后关键结果继续对齐，确认恢复动作是否真正生效？", "是否继续筛出残留影响对象，决定下一步是否需要追加恢复动作？", "要不要先确认恢复后的稳定窗口，再决定是否继续执行新的抢通动作？"], ["{service} 下一步怎么做？", "恢复动作是否生效？", "还有哪些对象未恢复？", "是否需要继续抢通？"], ["next_step", "recovery_fit"]),
    Blueprint("fault_guard", "故障处理", 10, "故障恢复后，现在在做 {service}，但还不知道下一步监控问题怎么设计更稳妥。", "围绕 {service} 的恢复后值守监控结果继续设计下一步问题，确认恢复状态和稳定性", ["确认恢复对象和监控窗口", "查看恢复后关键指标", "识别残留告警或波动", "判断是否已稳定", "决定是否继续观察"], ["当前已完成恢复后第一轮监控，发现多数对象已恢复，但仍有少量 {objects} 出现波动或残留告警。", "系统已拉取恢复后的关键指标，确认整体趋势向好，但局部对象仍未完全稳定。"], ["但还没有继续确认波动是否会复发，也没有明确哪些对象需要重点盯防。", "不过目前仍不清楚应继续看恢复状态、残留告警还是稳定窗口长度。"], ["是否继续按对象确认哪些恢复后对象仍存在残留波动或残留告警？", "要不要把恢复后的关键结果按时间窗口继续对齐，确认是否已经稳定？", "是否继续筛出需要重点值守的对象，避免把已稳定对象重复纳入？", "要不要先明确结束值守的判断标准，再继续确认恢复状态？"], ["恢复后怎么继续监控？", "还有残留告警吗？", "恢复状态稳定了吗？", "值守什么时候结束？"], ["next_step", "guard_fit"]),
    Blueprint("fault_ticket", "故障处理", 10, "{objects} 已经生成工单，现在只拿到一轮处理结果，下一步应该问什么？", "围绕工单处理结果继续设计下一步问题，确认工单对象、异常范围和后续动作", ["确认工单对象和异常来源", "核查关键结果和信号", "识别高影响对象", "判断下一步处理方向", "补充证据和恢复确认"], ["当前已完成工单初查，确认部分 {objects} 与 {alarm} 或关键异常信号直接相关。", "系统已返回工单处理结果，并识别出局部高影响对象和关键波动结果。"], ["但还没有继续按对象、时间或影响范围把工单问题进一步收敛。", "不过目前仍不清楚高影响对象是否集中，还是需要分别处理。"], ["是否继续按工单对象或影响范围确认问题是否集中？", "要不要把关键结果和时间窗口继续对齐，确认异常变化先后关系？", "是否继续筛出最需要优先处理的高影响对象？", "要不要先补足关键证据，再决定下一步是继续诊断还是直接推进恢复？"], ["工单下一步怎么处理？", "工单影响范围多大？", "高影响对象有哪些？", "下一步先诊断还是先恢复？"], ["next_step", "ticket_handling"]),
]


def pick(seq: list[str], idx: int) -> str:
    return seq[idx % len(seq)]


def slice_plan(steps: list[str], difficulty: str) -> list[str]:
    return steps[:3] if difficulty == "easy" else (steps[:4] if difficulty == "medium" else steps[:5])


def build_intent_value(primary: str, case_id: int) -> str | list[str]:
    if case_id % 8 != 0:
        return primary
    return {"知识问答": ["知识问答", "故障处理"], "网络监控": ["网络监控", "故障处理"], "故障处理": ["故障处理", "网络监控"]}[primary]


def role_for_intent(intent: str, idx: int) -> str:
    pools = {"知识问答": ["网络监控工程师", "网络规划工程师", "网络维护工程师"], "网络监控": ["网络监控工程师", "网络规划工程师", "网络监控工程师"], "故障处理": ["网络维护工程师", "网络监控工程师", "网络维护工程师"]}
    return pick(pools[intent], idx)


def build_user_profile(idx: int, intent: str, threshold: str) -> list[str]:
    role = role_for_intent(intent, idx)
    profile = [f"角色：{role}"]
    if idx % 2 == 0:
        profile.append(pick(GEO_TERMS, idx))
    if idx % 3 != 0:
        profile.append(pick(NETWORK_TERMS, idx))
    if idx % 4 != 0:
        profile.append(pick(BUSINESS_TERMS, idx))
    if intent == "网络监控":
        profile.append(f"异常判断偏好：超过 {threshold} 的波动视为显著异常")
    return profile


def build_execution_result(bp: Blueprint, plan: list[str], slots: dict[str, str], difficulty: str, idx: int) -> str:
    steps = "、".join(plan[: min(3, len(plan))])
    finding = pick(bp.findings, idx).format(**slots).rstrip("。")
    uncertainty = pick(bp.uncertainties, idx).format(**slots).rstrip("。")
    tail = {"easy": "当前线索相对集中，适合围绕主线继续推进。", "medium": "当前已有初步方向，但还缺一到两项关键确认动作。", "hard": "当前存在多个可能方向，下一步需要更强的范围收敛或证据补强。"}[difficulty]
    return f"已完成动作：{steps}。当前发现：{finding}。尚未解决：{uncertainty}。补充判断：{tail}"


def build_history(templates: list[str], slots: dict[str, str], idx: int) -> list[str]:
    if idx % 6 == 0:
        return []
    return [pick(templates, idx + j).format(**slots) for j in range(3 + (idx % 2))]


def build_next_top3(templates: list[str], slots: dict[str, str], idx: int) -> list[str]:
    out: list[str] = []
    for j in range(len(templates)):
        value = pick(templates, idx + j).format(**slots)
        if value not in out:
            out.append(value)
        if len(out) == 3:
            return out
    raise ValueError("Need 3 distinct next questions")


def render_case(case_id: int, bp: Blueprint, variant: int, difficulty: str) -> dict:
    pool = POOLS[bp.intent]
    region = pick(REGIONS, case_id + variant)
    threshold = pick(THRESHOLDS, case_id + variant)
    slots = {
        "region": region,
        "threshold": threshold,
        "site": pick(pool["sites"], case_id + variant),
        "time_window": pick(TIME_WINDOWS, case_id + variant),
        "node": pick(pool["nodes"], case_id + variant),
        "metrics": pick(pool["metrics"], case_id + variant),
        "alarm": pick(pool["alarms"], case_id + variant),
        "code": pick(pool["codes"], case_id + variant),
        "service": pick(pool["services"], case_id + variant),
        "objects": pick(pool["objects"], case_id + variant),
    }
    intent_value = build_intent_value(bp.intent, case_id)
    plan = slice_plan(bp.plan_steps, difficulty)
    execution_result = build_execution_result(bp, plan, slots, difficulty, case_id + variant)
    return {
        "id": f"NQ-{case_id:03d}",
        "intent": intent_value,
        "difficulty": difficulty,
        "input": {
            "original_question": bp.question_template.format(**slots),
            "rewritten_question": bp.rewritten_template.format(**slots),
            "intent": intent_value,
            "user_profile": build_user_profile(case_id + variant, bp.intent, threshold),
            "current_plan": plan,
            "execution_result": execution_result,
            "history_high_freq_questions": build_history(bp.history_templates, slots, case_id + variant),
        },
        "expected": {"top3": build_next_top3(bp.next_questions, slots, case_id + variant), "evaluation_focus": bp.evaluation_focus},
        "meta": {"scenario_family": bp.key, "region": region, "intent_mode": "multi" if isinstance(intent_value, list) else "single"},
    }


def build_cases() -> Iterable[dict]:
    case_id = 1
    diff_iter = iter(DIFFICULTIES)
    for bp in BLUEPRINTS:
        for variant in range(bp.repeats):
            yield render_case(case_id, bp, variant, next(diff_iter))
            case_id += 1


def serialize_excel_list(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value if str(item).strip())
    return str(value)


def parse_excel_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]

    text = str(value).strip()
    if not text:
        return []

    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]

    return [line.strip() for line in text.splitlines() if line.strip()]


def clean_cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def normalize_intent_from_excel(value: Any) -> str | list[str]:
    values = parse_excel_list(value)
    if not values:
        return ""
    return values if len(values) > 1 else values[0]


def case_to_excel_row(case: dict[str, Any]) -> list[Any]:
    return [
        case["id"],
        case["input"]["original_question"],
        case["input"]["rewritten_question"],
        serialize_excel_list(case["intent"]),
        case["difficulty"],
        serialize_excel_list(case["input"].get("user_profile") or []),
        serialize_excel_list(case["input"].get("current_plan") or []),
        case["input"].get("execution_result", ""),
        serialize_excel_list(case["input"].get("history_high_freq_questions") or []),
        serialize_excel_list(case["expected"].get("top3") or []),
        serialize_excel_list(case["expected"].get("evaluation_focus") or []),
        case["meta"].get("scenario_family", ""),
        case["meta"].get("region", ""),
        case["meta"].get("intent_mode", ""),
    ]


def row_to_case(row: dict[str, Any]) -> dict[str, Any]:
    intent_value = normalize_intent_from_excel(row.get("intent"))
    return {
        "id": clean_cell_text(row.get("编号") or row.get("id")),
        "intent": intent_value,
        "difficulty": clean_cell_text(row.get("difficulty")),
        "input": {
            "original_question": clean_cell_text(row.get("original_question")),
            "rewritten_question": clean_cell_text(row.get("rewritten_question")),
            "intent": intent_value,
            "user_profile": parse_excel_list(row.get("user_profile")),
            "current_plan": parse_excel_list(row.get("current_plan")),
            "execution_result": clean_cell_text(row.get("execution_result")),
            "history_high_freq_questions": parse_excel_list(row.get("history_high_freq_questions")),
        },
        "expected": {
            "top3": parse_excel_list(row.get("reference_top3")),
            "evaluation_focus": parse_excel_list(row.get("evaluation_focus")),
        },
        "meta": {
            "scenario_family": clean_cell_text(row.get("scenario_family")),
            "region": clean_cell_text(row.get("region")),
            "intent_mode": clean_cell_text(row.get("intent_mode")) or ("multi" if isinstance(intent_value, list) else "single"),
        },
    }


def write_excel(cases: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "评测集"
    worksheet.append(EXCEL_HEADERS)
    for case in cases:
        worksheet.append(case_to_excel_row(case))
    workbook.save(output_path)


def load_excel_dataset(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows, [])]
    cases: list[dict[str, Any]] = []
    for values in rows:
        row = dict(zip(headers, values, strict=False))
        if not any(value is not None and str(value).strip() for value in values):
            continue
        cases.append(row_to_case(row))
    return cases


def load_dataset(path: str | Path) -> list[dict[str, Any]]:
    dataset_path = Path(path)
    if dataset_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx dataset files are supported now: {dataset_path}")
    return load_excel_dataset(dataset_path)


def summarize_cases(cases: list[dict]) -> dict[str, object]:
    intents = Counter("+".join(x["intent"]) if isinstance(x["intent"], list) else x["intent"] for x in cases)
    diff = Counter(x["difficulty"] for x in cases)
    return {
        "count": len(cases),
        "intent_distribution": dict(intents),
        "difficulty_distribution": dict(diff),
        "multi_intent_cases": sum(isinstance(x["intent"], list) for x in cases),
    }


def print_summary(summary: dict[str, object]) -> None:
    print(f"Generated {summary['count']} cases")
    print(f"Intent distribution: {summary['intent_distribution']}")
    print(f"Difficulty distribution: {summary['difficulty_distribution']}")
    print(f"Multi-intent cases: {summary['multi_intent_cases']}")


def generate_dataset(output: str | Path = DEFAULT_DATASET_PATH) -> dict[str, object]:
    cases = list(build_cases())
    if len(cases) != TOTAL_CASES:
        raise ValueError(f"Expected {TOTAL_CASES} cases, got {len(cases)}")
    output_path = Path(output)
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx output files are supported now: {output_path}")
    write_excel(cases, output_path)
    summary = summarize_cases(cases)
    summary["output"] = str(output_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate telecom next-step-question evaluation set v3.")
    parser.add_argument("--output", default=DEFAULT_DATASET_PATH)
    args = parser.parse_args()
    summary = generate_dataset(args.output)
    print_summary(summary)


if __name__ == "__main__":
    main()
